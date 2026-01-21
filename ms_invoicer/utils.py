import json
from typing import Optional, Union
import logging
import os
import boto3
import pytz

from typing import List
from fastapi import UploadFile
from datetime import datetime, time, timedelta
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

from ms_invoicer.config import S3_ACCESS_KEY, S3_SECRET_ACCESS_KEY

log = logging.getLogger(__name__)


MONTH_NAMES_FRENCH = {
    1: 'Janvier',
    2: 'Février',
    3: 'Mars',
    4: 'Avril',
    5: 'Mai',
    6: 'Juin',
    7: 'Juillet',
    8: 'Août',
    9: 'Septembre',
    10: 'Octobre',
    11: 'Novembre',
    12: 'Décembre'
}


class BreadCrumbs:
    def __init__(self, href: str, value: str, required_id: int = None):
        self.href = href
        self.value = value
        self.required_id = required_id
    
    def __repr__(self):
        return json.dumps({"href": self.href, "value": self.value, "required_id": self.required_id})


def check_dates(base_date: datetime, date1: time, date2: time = None) -> datetime:
    """ """
    result = datetime.combine(base_date, date1)
    if date2 and date1 < date2:
        result = result + timedelta(days=1)
    return result


def create_folders():
    folder_names = ["temp", "temp/xlsx", "temp/pdf", "temp/cache"]
    for folder_name in folder_names:
        if not os.path.exists(folder_name):
            log.info(
                "Created folder",
                extra={"folder_path": folder_name, "event": "create_folder"},
            )
            os.mkdir(folder_name)
        else:
            log.debug(
                "Folder already exists",
                extra={"folder_path": folder_name, "event": "create_folder"},
            )


def save_file(file_path: str, file: UploadFile):
    if os.path.exists(file_path):
        os.remove(file_path)
    with open(file_path, "wb") as output_file:
        output_file.write(file.file.read())


def remove_file(file_path: str):
    if os.path.isfile(file_path):
        os.remove(file_path)
    else:
        log.warning(
            "File not found for removal",
            extra={"file_path": file_path, "event": "remove_file"},
        )


def find_ranges(sheet: Worksheet) -> List[tuple]:
    """
    This function takes an Excel worksheet (using the openpyxl library) as input,
    searches for a specific string in the first column of the sheet, and returns
    a list of tuples representing the row ranges that contain the string.

    Input:

    sheet: an Excel worksheet object (type: Worksheet)

    Output:

    A list of tuples, where each tuple contains four integers representing the
    starting and ending rows of a range that contains the target string.

    Functionality:

    The function iterates through each row of the given worksheet, stopping at
    row 200 and column 10. For each row, it checks the value of the first cell
    (column A). If the value is a non-empty string containing the text "NOM CONTRAT",
    the function records the starting row number and adds a tuple to the result list.
    The tuple contains the starting row number, the row number 2 rows below the
    starting row, the row number 4 rows below the starting row, and the row number
    34 rows below the starting row. These row numbers are calculated based on the
    assumption that each contract occupies 31 rows (31 days in a month).
    """
    result = []
    start_num_info = None
    start_num = None
    end_num = None
    for row in sheet.iter_rows(max_col=10, max_row=200):
        row_data: Cell = row[0]
        if row_data.value and isinstance(row_data.value, str):
            if (
                row_data.value.lower().startswith("nom contrat") 
                or row_data.value.lower().startswith("nom")
            ):
                start_num_info = int(row_data.row)
            elif row_data.value.lower().startswith("total"):
                end_num = int(row_data.row)
            elif row_data.value.lower().startswith("date"):
                start_num = int(row_data.row) + 1

        if start_num_info and end_num and start_num:
            result.append((start_num_info, start_num_info + 2, start_num, end_num))
            start_num_info = None
            end_num = None
    return result


def find_letter(sheet: Worksheet) -> List[tuple]:
    result = None
    for row in sheet.iter_rows(max_col=10, max_row=200):
        row_data: Cell = row[0]
        if row_data.value and isinstance(row_data.value, str):
            if row_data.value.lower().startswith("date"):
                    for col in sheet.iter_cols(min_row=int(row_data.row), max_row=int(row_data.row)+1, max_col=30):
                        col_data: Cell = col[0]
                        if (
                            col_data.value
                            and isinstance(col_data.value, str)
                            and col_data.value.lower().startswith("heures")
                        ):
                            result = col_data.column_letter
                            break
        if result:
            break
    return result


def upload_file(
    file_path: str,
    file_name: str,
    bucket="invoicer-dev-01",
    object_name=None,
    is_pdf=False,
):
    """Upload a file to an S3 bucket

    :param file_name: File to upload
    :param bucket: Bucket to upload to
    :param object_name: S3 object name. If not specified then file_name is used
    :return: True if file was uploaded, else False
    """

    """if ENV != "production":
        url = "https://" + bucket + ".s3.amazonaws.com/" + file_name
        return url"""

    # If S3 object_name was not specified, use file_name
    if object_name is None:
        object_name = os.path.basename(file_path)

    # Upload the file
    session: boto3.Session = boto3.session.Session()
    s3 = session.client(
        "s3",
        aws_access_key_id=S3_ACCESS_KEY,
        aws_secret_access_key=S3_SECRET_ACCESS_KEY,
    )
    try:
        args = {"ACL": "public-read"}
        if is_pdf:
            try:
                download_filename = file_name.split("-")[0]
            except:
                download_filename = file_name
            args = {
                "ACL": "public-read",
                "ContentType": "application/pdf",
                "ContentDisposition": 'inline; filename="{}.pdf"'.format(
                    download_filename
                ),
            }
        s3.upload_file(
            file_path,
            bucket,
            object_name,
            ExtraArgs=args,
        )
        url = "https://" + bucket + ".s3.amazonaws.com/" + file_name
        return url
    except Exception as e:
        log.exception(
            "Failed to upload file to S3",
            extra={
                "bucket": bucket,
                "file_path": file_path,
                "object_name": object_name,
                "is_pdf": is_pdf,
                "event": "upload_file",
            },
        )
        raise Exception("Failure uploading file")


def download_file_from_s3(
    file_path_s3: str,
    path_to_save: str,
    bucket="invoicer-dev-01",
):
    try:
        session: boto3.Session = boto3.session.Session()
        s3 = session.client(
            "s3",
            aws_access_key_id=S3_ACCESS_KEY,
            aws_secret_access_key=S3_SECRET_ACCESS_KEY,
        )

        # Download the file
        s3.download_file(bucket, file_path_s3, path_to_save)
        if not os.path.exists(path_to_save):
            raise Exception("Failure downloading file")
    except Exception as e:
        log.exception(
            "Failed to download file from S3",
            extra={
                "bucket": bucket,
                "file_path_s3": file_path_s3,
                "path_to_save": path_to_save,
                "event": "download_file",
            },
        )
        raise Exception("Failure downloading file")


def delete_file_from_s3(file_names: Union[str, list], bucket_name="invoicer-dev-01"):
    try:
        # Create a session using your AWS credentials
        s3_client = boto3.client(
            "s3",
            aws_access_key_id=S3_ACCESS_KEY,
            aws_secret_access_key=S3_SECRET_ACCESS_KEY,
        )
        if isinstance(file_names, list):
            for file_name in file_names:
                s3_client.delete_object(Bucket=bucket_name, Key=file_name)
        else:
            s3_client.delete_object(Bucket=bucket_name, Key=file_names)
    except Exception as e:
        log.exception(
            "Failed to delete file(s) from S3",
            extra={
                "bucket": bucket_name,
                "file_names": file_names,
                "event": "delete_file",
            },
        )
        raise Exception("Failure deleting file")


def extract_and_get_month_name(date: datetime) -> Optional[str]:
    month = date.month
    if month in MONTH_NAMES_FRENCH:
        return MONTH_NAMES_FRENCH[month]
    return None


def get_current_date() -> datetime:
    # Create a timezone object for Eastern Standard Time
    est = pytz.timezone('US/Eastern')
    # Get the current time in UTC and convert it to EST
    current_time_in_est = datetime.now(pytz.utc).astimezone(est)
    return current_time_in_est
