import os
import bs4
import jinja2
import pdfkit
import logging

from datetime import datetime
from uuid import uuid4
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from pypdf import PdfMerger
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import PageBreak, SimpleDocTemplate, Table, TableStyle

from ms_invoicer.config import S3_BUCKET_NAME, WKHTMLTOPDF_PATH
from ms_invoicer.dao import (
    GenerateFinalPDFNoFile,
    GenerateFinalPDFWithFile,
    PdfToProcessEvent,
)
from ms_invoicer.db_pool import get_db
from ms_invoicer.event_bus import publish
from ms_invoicer.utils import find_ranges, upload_file
from ms_invoicer.sql_app import crud

base = os.path.dirname(os.path.dirname(__file__))
log = logging.getLogger(__name__)


async def build_pdf(event: PdfToProcessEvent):
    try:
        log.info(
            "Customer {} - Initiated building pdf - fun: build_pdf".format(
                event.current_user_id
            )
        )
        connection = next(get_db())
        input_html_path: str = os.path.join(
            base, "templates/base/{}".format(event.html_template_name)
        )
        output_html_path: str = os.path.join(
            base, "templates/{}".format(event.html_template_name)
        )
        event.file = crud.get_file(
            db=connection, current_user_id=event.current_user_id, model_id=event.file.id
        )
        event.invoice = crud.get_invoice(
            db=connection,
            current_user_id=event.current_user_id,
            model_id=event.invoice.id,
        )
        with open(input_html_path) as fp:
            soup = BeautifulSoup(fp, "html.parser")

            new_tag = soup.new_tag("div")
            for index in range(0, len(event.file.services)):
                new_tr_tag = soup.new_tag("tr")
                for variable in [
                    "service_id",
                    "service_txt",
                    "num_hours",
                    "amount",
                ]:
                    new_td_tag = soup.new_tag("td")
                    new_td_tag.string = "{{" + variable + str(index) + "}}"
                    new_tr_tag.append(new_td_tag)
                new_tag.append(new_tr_tag)
            for comment in soup.find_all(string=lambda e: isinstance(e, bs4.Comment)):
                if "items" in comment:
                    comment.replace_with(new_tag)

            with open(output_html_path, "w") as file:
                file.write(str(soup))

            log.info(
                "Customer {} - Adding pdf data - fun: build_pdf".format(
                    event.current_user_id
                )
            )
            top_info_data = {}
            top_info = crud.get_topinfos(
                db=connection, current_user_id=event.current_user_id
            )
            if len(top_info) > 0:
                top_info = top_info[0]
                top_info_data["top_info_from"] = top_info.ti_from
                top_info_data["top_info_addr"] = top_info.addr
                top_info_data["top_info_phone"] = top_info.phone
                top_info_data["top_info_email"] = top_info.email

            bill_to_data = {}
            if event.file.bill_to:
                bill_to_data["to"] = event.file.bill_to.to
                bill_to_data["addr"] = event.file.bill_to.addr
                bill_to_data["phone"] = event.file.bill_to.phone
                bill_to_data["email"] = event.file.bill_to.email

            service_data = {}
            subtotal = 0
            index = 0
            service_info = None
            # Contratos enviados sin archivo
            for service in event.file.services:
                if not service_info:
                    service_info = service
                service_data["service_id{}".format(index)] = index + 1
                service_data["service_txt{}".format(index)] = service.title
                service_data["num_hours{}".format(index)] = service.hours
                service_data["amount{}".format(index)] = service.amount
                subtotal += service.amount
                index += 1

            if event.invoice.with_taxes:
                total_tax_1 = (event.invoice.tax_1 / 100) * subtotal
                total_tax_2 = (event.invoice.tax_2 / 100) * subtotal
                total = total_tax_1 + total_tax_2 + subtotal
            else:
                total = subtotal

            title_company = "-"
            empresa_variable = crud.get_global(
                db=connection, identifier=3, current_user_id=event.current_user_id
            )
            if empresa_variable:
                title_company = empresa_variable.value

            variables = crud.get_globals(
                db=connection, current_user_id=event.current_user_id
            )
            tps_name = "TPS -"
            tvq_name = "TVQ -"
            for variable in variables:
                if variable.identifier == 1:
                    tps_name = variable.name
                elif variable.identifier == 2:
                    tvq_name = variable.name

            context = {
                "title_company": title_company,
                "invoice_date": event.invoice.created.strftime("%d-%m-%Y"),
                "invoice_id": event.invoice.number_id,
                "created": event.invoice.created,
                "total": round(total, 2),
            }

            if event.invoice.with_taxes:
                context["tps_name"] = tps_name
                context["tvq_name"] = tvq_name
                context["total_no_taxes"] = round(subtotal, 2)
                context["total_tax1"] = round(total_tax_1, 2)
                context["total_tax2"] = round(total_tax_2, 2)

            context |= bill_to_data
            context |= service_data
            context |= top_info_data

            log.info(
                "Customer {} - Creating pdf - fun: build_pdf".format(
                    event.current_user_id
                )
            )
            template_loader = jinja2.FileSystemLoader(os.path.join(base, "templates"))
            template_env = jinja2.Environment(loader=template_loader)

            template = template_env.get_template(event.html_template_name)
            output_text = template.render(context)

            if not service_info:
                searched_file = crud.get_file(
                    db=connection,
                    model_id=event.file.id,
                    current_user_id=event.current_user_id,
                )
                if searched_file and len(searched_file.services) > 0:
                    service_info = searched_file.services[0]
            filename = f"facture_{event.invoice.number_id}_{service_info.title}_{event.invoice.created.strftime('%m_%Y')}-{str(uuid4())}.pdf"
            filename = filename.replace(" ", "_")
            output_pdf_path: str = "temp/pdf/{}".format(filename)
            config = pdfkit.configuration(
                wkhtmltopdf=WKHTMLTOPDF_PATH
            )  # TODO: Modificar en la máquina
            pdfkit.from_string(output_text, output_pdf_path, configuration=config)

            if event.with_file:
                data_event = GenerateFinalPDFWithFile(
                    current_user_id=event.current_user_id,
                    pdf_tables=f"temp/pdf_tables_{event.current_user_id}.pdf",
                    xlsx_url=event.xlsx_url,
                    pdf_invoice=output_pdf_path,
                    filename=filename,
                    file_id=event.file.id,
                    with_tables=event.invoice.with_tables,
                    pages=event.pages,
                )
            else:
                data_event = GenerateFinalPDFNoFile(
                    current_user_id=event.current_user_id,
                    pdf_invoice=output_pdf_path,
                    filename=filename,
                    file_id=event.file.id,
                )
            log.info(
                "Customer {} - Publishing GenerateFinalPDF event - fun: build_pdf".format(
                    event.current_user_id
                )
            )
            await publish(data_event)
            return True
    except Exception as e:
        log.error("Customer {} - Failure building pdf".format(event.current_user_id))
        log.error(e)
        raise


def generate_invoice(event: GenerateFinalPDFWithFile):
    log.info(
        "Customer {} - Building invoice - fun: generate_invoice".format(
            event.current_user_id
        )
    )
    try:
        paths_to_check = [event.path_pdf_invoice]
        if event.with_tables:
            paths_to_check.append(event.path_pdf_tables)
            # Define the page size
            page_size = letter

            # Define the table style
            table_style = TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#333333")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#ffffff")),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("LEADING", (0, 0), (-1, -1), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f2f2f2")),
                    ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#333333")),
                    ("GRID", (0, 0), (-1, 0), 2, colors.HexColor("#333333")),
                    ("GRID", (0, 1), (-1, -1), 1, colors.HexColor("#333333")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )

            # Create the PDF document
            doc = SimpleDocTemplate(event.path_pdf_tables, pagesize=page_size)
            elements = []

            # Load the Excel workbook
            wb = load_workbook(filename=event.xlsx_url, data_only=True)

            # Select the worksheet with the table
            for sheet_name in wb.sheetnames:
                if sheet_name not in event.pages:
                    continue
                ws = wb[sheet_name]

                ranges = find_ranges(ws)
                for range_of in ranges:
                    (_, _, minrow, maxrow) = range_of

                    # -1 because the headers
                    # Extract the table data as a list of rows
                    table_data = []
                    have_nom = False
                    for row in ws.iter_rows(
                        min_row=minrow - 1, max_row=maxrow, max_col=7, values_only=True
                    ):
                        row_data = []
                        for index, cell in enumerate(row):
                            if cell and isinstance(cell, str) and ("Nom, Prenom" in cell or "nom, prenom" in cell):
                                have_nom = True
                            if index == 1 and have_nom:
                                continue
                            if cell is None and index == 6:
                                continue
                            if isinstance(cell, datetime):
                                cell_value = cell.strftime("%Y-%m-%d")
                            elif isinstance(cell, float):
                                cell_value = round(cell, 2)
                            else:
                                cell_value = cell
                            row_data.append(cell_value)
                        table_data.append(row_data)

                    # Add the table to the PDF document
                    table = Table(table_data)
                    table.setStyle(table_style)
                    elements.append(table)

                    # Add a page break after the table
                    elements.append(PageBreak())

            # Build and save the PDF document
            doc.build(elements)

        merger = PdfMerger()
        for pdf in paths_to_check:
            merger.append(pdf)
        merger.write(event.path_pdf_invoice)
        merger.close()

        log.info(
            "Customer {} - Uploading invoice - fun: generate_invoice".format(
                event.current_user_id
            )
        )
        s3_pdf_url = upload_file(
            file_path=event.path_pdf_invoice,
            file_name=event.filename,
            is_pdf=True,
            bucket=S3_BUCKET_NAME,
        )
        conn = next(get_db())
        crud.patch_file(
            db=conn,
            model_id=event.file_id,
            update_dict={"s3_pdf_url": s3_pdf_url},
            current_user_id=event.current_user_id,
        )
        log.info(
            "Customer {} - Invoice generated - fun: generate_invoice".format(
                event.current_user_id
            )
        )
        return True
    except Exception as e:
        log.error(
            "Customer {} - Failure generating invoice".format(event.current_user_id)
        )
        log.error(e)
        raise


def generate_invoice_no_file(event: GenerateFinalPDFNoFile):
    log.info(
        "Customer {} - Building invoice - fun: generate_invoice_no_file".format(
            event.current_user_id
        )
    )
    try:
        log.info(
            "Customer {} - Uploading invoice - fun: generate_invoice_no_file".format(
                event.current_user_id
            )
        )
        s3_pdf_url = upload_file(
            file_path=event.path_pdf_invoice,
            file_name=event.filename,
            is_pdf=True,
            bucket=S3_BUCKET_NAME,
        )
        conn = next(get_db())
        crud.patch_file(
            db=conn,
            model_id=event.file_id,
            update_dict={"s3_pdf_url": s3_pdf_url},
            current_user_id=event.current_user_id,
        )
        log.info(
            "Customer {} - Invoice generated - fun: generate_invoice_no_file".format(
                event.current_user_id
            )
        )
        return True
    except Exception as e:
        log.error(
            "Customer {} - Failure generating invoice".format(event.current_user_id)
        )
        log.error(e)
        raise
