from ctypes import Union
import logging
import shutil
import os
from typing import Dict, List, Tuple, Union
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from datetime import datetime
from pathlib import Path
from uuid import uuid4

from fastapi import UploadFile, status, HTTPException
from sqlalchemy.orm import Session
from openpyxl.cell.cell import Cell
from ms_invoicer.config import S3_BUCKET_NAME

from ms_invoicer.dao import FilesToProcessEvent, PdfToProcessEvent
from ms_invoicer.db_pool import get_db_context
from ms_invoicer.event_bus import publish
from ms_invoicer.sql_app import crud, schemas
from ms_invoicer.sql_app import models
from ms_invoicer.sql_app.models import Invoice, User
from ms_invoicer.utils import (
    download_file_from_s3,
    extract_and_get_month_name,
    find_letter,
    remove_file,
    save_file,
    find_ranges,
    upload_file,
    get_current_date,
)

log = logging.getLogger(__name__)


def extract_pages(uploaded_file: UploadFile, current_user_id: int) -> List[str]:
    log.info(
        "Extracting xlsx pages",
        extra={"customer_id": current_user_id, "event": "extract_pages"},
    )
    try:
        date_now = get_current_date()

        filename = f"{date_now.year}{date_now.month}{date_now.day}{date_now.hour}{date_now.minute}{date_now.second}-{str(uuid4())}.xlsx"
        filename = filename.replace(" ", "_")
        file_path = "temp/xlsx/{}".format(filename)
        save_file(file_path, uploaded_file)

        xlsx_file = Path(file_path)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        results = wb_obj.sheetnames
        remove_file(file_path)
        return results
    except Exception:
        log.exception(
            "Failed to extract xlsx pages",
            extra={"customer_id": current_user_id, "event": "extract_pages"},
        )
        return []


async def process_file(
    file: Union[UploadFile, None],
    invoice_id: int,
    bill_to_id: int,
    current_user_id: int,
    col_letter: str = "F",
    pages: Union[List[str], None] = None,
) -> schemas.File:
    if file is None:
        return None, None
    pages = pages or []
    try:
        log.info(
            "Processing xlsx upload",
            extra={
                "customer_id": current_user_id,
                "invoice_id": invoice_id,
                "bill_to_id": bill_to_id,
                "event": "process_file",
            },
        )
        date_now = get_current_date()
        filename = f"{date_now.year}{date_now.month}{date_now.day}{date_now.hour}{date_now.minute}{date_now.second}-{str(uuid4())}.xlsx"
        filename = filename.replace(" ", "_")
        file_path = "temp/xlsx/{}".format(filename)
        save_file(file_path, file)

        price_unit = 1
        currency = "CAD"
        log.debug(
            "Uploading xlsx to S3",
            extra={
                "customer_id": current_user_id,
                "invoice_id": invoice_id,
                "event": "process_file",
            },
        )
        s3_url = upload_file(
            file_path=file_path, file_name=filename, bucket=S3_BUCKET_NAME
        )
        file_obj = schemas.FileCreate(
            **{
                "s3_xlsx_url": s3_url,
                "s3_pdf_url": None,
                "created": date_now,
                "pages_xlsx": ",".join(pages),
                "invoice_id": invoice_id,
                "bill_to_id": bill_to_id,
                "user_id": current_user_id,
            }
        )
        with get_db_context() as db:
            new_file = crud.create_file(db=db, model=file_obj)
            data_event = FilesToProcessEvent(
                file_path=file_path,
                file_id=new_file.id,
                invoice_id=invoice_id,
                current_user_id=current_user_id,
                price_unit=price_unit,
                col_letter=col_letter,
                currency=currency,
                pages=pages,
            )
            log.info(
                "Publishing file processing event",
                extra={
                    "customer_id": current_user_id,
                    "invoice_id": invoice_id,
                    "file_id": new_file.id,
                    "event": "process_file",
                },
            )
            await publish(data_event)
            return new_file, file_path
    except Exception:
        log.exception(
            "Failed to process xlsx upload",
            extra={
                "customer_id": current_user_id,
                "invoice_id": invoice_id,
                "bill_to_id": bill_to_id,
                "event": "process_file",
            },
        )
        raise


async def extract_data(event: FilesToProcessEvent) -> bool:
    log.info(
        "Extracting invoice data from xlsx",
        extra={
            "customer_id": event.current_user_id,
            "invoice_id": event.invoice_id,
            "file_id": event.file_id,
            "event": "extract_data",
        },
    )
    event_handled = False
    with get_db_context() as conn:
        try:
            currency = event.currency

            xlsx_file = Path(event.xlsx_url)
            wb_obj = openpyxl.load_workbook(xlsx_file, data_only=True)
            for sheet_name in wb_obj.sheetnames:
                if sheet_name not in event.pages:
                    continue
                sheet: Cell = wb_obj[sheet_name]
                col_letter = find_letter(sheet)

                ranges = find_ranges(sheet)
                invoice_update = False
                for range_of in ranges:
                    (minrowinfo, maxrowinfo, minrow, maxrow) = range_of
                    if not invoice_update:
                        parsed_date = str(sheet["{}{}".format("A", maxrow - 1)].value)
                        date_formats = [
                            "%d/%m/%Y %H:%M:%S",
                            "%d/%m/%Y",
                            "%d-%m-%Y %H:%M:%S",
                            "%d-%m-%Y",
                        ]
                        for date_format in date_formats:
                            try:
                                parsed_date = datetime.strptime(parsed_date, date_format)
                                break  # Break out of the loop if parsing is successful
                            except ValueError:
                                pass

                        crud.patch_invoice(
                            db=conn,
                            model_id=event.invoice_id,
                            current_user_id=event.current_user_id,
                            update_dict={"created": parsed_date},
                        )
                        invoice_update = True
                    contract_dict = {}
                    for row in sheet.iter_rows(
                        min_row=minrowinfo, max_row=maxrowinfo, max_col=15
                    ):
                        cell = row[0].value
                        if cell:
                            if "NOM CONTRAT" in cell or "NOM" in cell:
                                contract_dict["title"] = row[2].value
                            elif "HEURE" in cell or "HEURES" in cell:
                                contract_dict["price_unit"] = row[2].value
                            elif "MONTANT" in cell:
                                contract_dict["total_amount"] = row[2].value
                    hours: float = 0
                    for col in range(minrow, maxrow):
                        hour: float = sheet["{}{}".format(col_letter, col)].value
                        hours += float(round(hour, 2))

                    price_unit = contract_dict.get("price_unit", None)
                    amount = 0
                    if not price_unit:
                        amount = float(contract_dict.get("total_amount", 1))
                        price_unit = round(amount / hours, 2)
                    else:
                        str_price_unit = str(price_unit).replace("$", "").strip()
                        price_unit = float(str_price_unit)
                        amount = hours * float(price_unit)
                    contract_dict["hours"] = hours
                    contract_dict["currency"] = currency
                    contract_dict["amount"] = round(amount, 2)
                    contract_dict["price_unit"] = price_unit
                    contract_dict["file_id"] = event.file_id
                    contract_dict["invoice_id"] = event.invoice_id
                    contract_dict["user_id"] = event.current_user_id
                    service_created = schemas.ServiceCreate(**contract_dict)
                    crud.create_service(db=conn, model=service_created)

            log.info(
                "Extracted invoice data from xlsx",
                extra={
                    "customer_id": event.current_user_id,
                    "invoice_id": event.invoice_id,
                    "file_id": event.file_id,
                    "event": "extract_data",
                },
            )
            event_handled = True
            return event_handled
        except Exception:
            log.exception(
                "Failed to extract invoice data from xlsx",
                extra={
                    "customer_id": event.current_user_id,
                    "invoice_id": event.invoice_id,
                    "file_id": event.file_id,
                    "event": "extract_data",
                },
            )
            raise


async def process_pdf(
    db: Session,
    invoice: Invoice,
    bill_to_id: int,
    contracts: List[schemas.ServiceCreateNoFile],
    current_user: User,
    new_file_obj: schemas.File,
    xlsx_local_path: str,
    pages: Union[List[str], None] = None,
):
    pages = pages or []
    log.info(
        "Processing PDF creation",
        extra={
            "customer_id": current_user.id,
            "invoice_id": invoice.id,
            "bill_to_id": bill_to_id,
            "event": "process_pdf",
        },
    )
    current_date = get_current_date()
    with_file = False
    if not new_file_obj:
        file_obj = schemas.FileCreate(
            **{
                "s3_xlsx_url": None,
                "s3_pdf_url": None,
                "created": current_date,
                "pages_xlsx": ",".join(pages),
                "invoice_id": invoice.id,
                "bill_to_id": bill_to_id,
                "user_id": current_user.id,
            }
        )
        new_file = crud.create_file(db=db, model=file_obj)
        log.info(
            "Created file record without xlsx",
            extra={
                "customer_id": current_user.id,
                "invoice_id": invoice.id,
                "file_id": new_file.id,
                "event": "process_pdf",
            },
        )
    else:
        new_file = new_file_obj
        with_file = True
        log.debug(
            "Using existing xlsx for PDF creation",
            extra={
                "customer_id": current_user.id,
                "invoice_id": invoice.id,
                "file_id": new_file.id,
                "event": "process_pdf",
            },
        )

    result = []
    for contract in contracts:
        obj_dict = contract.model_dump()
        obj_dict["user_id"] = current_user.id
        obj_dict["invoice_id"] = invoice.id
        obj_dict["file_id"] = new_file.id
        result.append(
            crud.create_service(db=db, model=schemas.ServiceCreate(**obj_dict))
        )
    log.info(
        "Created invoice services",
        extra={
            "customer_id": current_user.id,
            "invoice_id": invoice.id,
            "file_id": new_file.id,
            "service_count": len(result),
            "event": "process_pdf",
        },
    )

    template_name = "template01.html"
    if not invoice.with_taxes:
        template_name = "template03.html"
    data_event = PdfToProcessEvent(
        current_user_id=current_user.id,
        invoice=invoice,
        file=new_file,
        html_template_name=template_name,
        xlsx_url=xlsx_local_path,
        with_file=with_file,
        pages=pages,
    )
    log.info(
        "Publishing PDF build event",
        extra={
            "customer_id": current_user.id,
            "invoice_id": invoice.id,
            "file_id": new_file.id,
            "event": "process_pdf",
        },
    )
    await publish(data_event)
    crud.patch_invoice(
        db=db,
        model_id=invoice.id,
        current_user_id=current_user.id,
        update_dict={"updated": current_date},
    )
    return crud.get_file(db=db, model_id=new_file.id, current_user_id=current_user.id)


async def generate_summary_by_date(
    db: Session,
    customer_id: int,
    current_user: User,
    start_date: datetime,
    end_date: datetime,
):
    log.info(
        "Generating summary by date",
        extra={
            "customer_id": current_user.id,
            "target_customer_id": customer_id,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "event": "generate_summary_by_date",
        },
    )
    invoices_by_customer = crud.get_invoices_by_customer_and_date_range(
        db=db,
        model_id=customer_id,
        current_user_id=current_user.id,
        start_date=start_date,
        end_date=end_date,
    )

    xlsx_list: List[models.File] = []
    for invoice in invoices_by_customer:
        files_list = crud.get_files_by_invoice(
            db=db, model_id=invoice.id, current_user_id=current_user.id
        )
        if len(files_list) > 0:
            xlsx_list.append(files_list[0])

    class FileToProcess:
        def __init__(
            self, filename: str, file_path: str, invoice_id: int, pages_xlsx: str
        ) -> None:
            self.filename = filename
            self.file_path = file_path
            self.invoice_id = invoice_id
            self.pages_xlsx = pages_xlsx

        def __str__(self) -> str:
            return "file_path: {} - invoice_id: {}".format(
                self.file_path, self.invoice_id, self.pages_xlsx
            )

    log.info(
        "Downloading xlsx files for summary",
        extra={
            "customer_id": current_user.id,
            "target_customer_id": customer_id,
            "event": "generate_summary_by_date",
        },
    )
    xlsx_path_name_list: List[FileToProcess] = []
    for xlsx in xlsx_list:
        filename = f"to_process_{xlsx.invoice_id}_{xlsx.id}_{xlsx.user_id}.xlsx"
        file_path = "temp/xlsx/{}".format(filename)
        new_file_to_process = FileToProcess(
            filename, file_path, xlsx.invoice_id, xlsx.pages_xlsx
        )
        xlsx_path_name_list.append(new_file_to_process)
        if not os.path.exists(file_path) and xlsx.s3_xlsx_url is not None:
            splitted_name = xlsx.s3_xlsx_url.split("/")
            download_file_from_s3(
                file_path_s3=splitted_name[3],
                path_to_save=file_path,
                bucket=S3_BUCKET_NAME,
            )
    xlsx_path_name_list = sorted(xlsx_path_name_list, key=lambda x: x.invoice_id)

    try:
        customer_obj = crud.get_customer(
            db=db, model_id=customer_id, current_user_id=current_user.id
        )
        output_filename = "resumen_{}_{}_{}-{}_{}.xlsx".format(
            customer_obj.name.replace(" ", "_"),
            start_date.month,
            start_date.year,
            end_date.month,
            end_date.year,
        )
        output_file_path = "temp/xlsx/{}".format(output_filename)
        shutil.copy("scripts/summary_template.xlsx", output_file_path)

        new_workbook = openpyxl.load_workbook(output_file_path)
        pair_sheet_names: Dict[str, List[Tuple]] = {}
        for path in xlsx_path_name_list:
            xlsx_file = Path(path.file_path)
            wb_obj = openpyxl.load_workbook(xlsx_file, data_only=True)

            if path.file_path not in pair_sheet_names:
                pair_sheet_names[path.file_path] = []
            pages_xlsx: List[str] = []
            should_check = False
            if path.pages_xlsx is not None and path.pages_xlsx != "":
                pages_xlsx = path.pages_xlsx.split(",")
                should_check = True
            for _, sheet_name in enumerate(wb_obj.sheetnames):
                sheet: Worksheet = wb_obj[sheet_name]
                if should_check and sheet_name.replace(",", "-") not in pages_xlsx:
                    continue
                new_sheet: Worksheet = new_workbook["01"]
                new_worksheet: Worksheet = new_workbook.copy_worksheet(new_sheet)
                pair_sheet_names[path.file_path].append(
                    (sheet.title, new_worksheet.title)
                )

        index_contract = 1
        for path in xlsx_path_name_list:
            xlsx_file = Path(path.file_path)
            wb_obj = openpyxl.load_workbook(xlsx_file, data_only=True)
            for input_sheet_name, output_sheet_name in pair_sheet_names[path.file_path]:
                input_sheet: Worksheet = wb_obj[input_sheet_name]

                new_sheet_wb: Worksheet = new_workbook[output_sheet_name]
                new_sheet_wb.title = "{}".format(index_contract)
                index_contract += 1

                ranges = find_ranges(input_sheet)
                period_extracted = False
                for range_of in ranges:
                    (minrowinfo, maxrowinfo, minrow, maxrow) = range_of
                    for row in input_sheet.iter_rows(
                        min_row=minrowinfo, max_row=maxrowinfo, max_col=10
                    ):
                        cell = row[0].value
                        if cell:
                            if "NOM CONTRAT" in cell:
                                cell_new_wb = new_sheet_wb["C1"]
                                cell_new_wb.value = row[2].value

                    if not period_extracted:
                        new_sheet_wb[
                            "{}{}".format("C", 2)
                        ].value = extract_and_get_month_name(
                            input_sheet["{}{}".format("A", minrow)].value
                        )
                        period_extracted = True

                    # Values
                    for col in range(minrow, maxrow):
                        input_date_value = input_sheet["{}{}".format("A", col)].value
                        new_sheet_wb["{}{}".format("A", col)].value = input_date_value

                        input_date_value = input_sheet["{}{}".format("B", col)].value
                        new_sheet_wb["{}{}".format("C", col)].value = input_date_value

                        input_date_value = input_sheet["{}{}".format("C", col)].value
                        new_sheet_wb["{}{}".format("D", col)].value = input_date_value

                        input_date_value = input_sheet["{}{}".format("D", col)].value
                        new_sheet_wb["{}{}".format("E", col)].value = input_date_value

                        input_date_value = input_sheet["{}{}".format("E", col)].value
                        new_sheet_wb["{}{}".format("F", col)].value = input_date_value

                        input_date_value = input_sheet["{}{}".format("F", col)].value
                        new_sheet_wb["{}{}".format("G", col)].value = input_date_value

                    if maxrow - minrow <= 30:
                        new_sheet_wb.delete_rows(36)
                        new_sheet_wb["{}{}".format("G", 36)].value = "=SUM(G6:G35)"
                        new_sheet_wb["{}{}".format("G", 38)].value = "=SUM(G22:G35)"
                        new_sheet_wb["{}{}".format("I", 36)].value = "=SUM(I6:I35)"
                        new_sheet_wb["{}{}".format("K", 36)].value = "=SUM(K6:K35)"

        original_sheet = new_workbook["01"]
        new_workbook.remove(original_sheet)
        new_workbook.save(output_file_path)
        # Close the workbook
        new_workbook.close()
        s3_url = upload_file(
            file_path=output_file_path, file_name=output_filename, bucket=S3_BUCKET_NAME
        )
        return s3_url

    except Exception:
        log.exception(
            "Failed to generate summary by date",
            extra={
                "customer_id": current_user.id,
                "target_customer_id": customer_id,
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "event": "generate_summary_by_date",
            },
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Hubo un error generanto el resumen",
        )
